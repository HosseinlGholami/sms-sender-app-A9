import openpyxl

FILE_NAME="./a.xlsx"

def get_number_from_excel_file(FILE_NAME):
    wb= openpyxl.load_workbook(FILE_NAME)
    sheet=wb[wb.sheetnames[0]]
    stop=False
    i=0
    numbers=list()
    while(stop!=True):
        i+=1
        numbers.append([sheet.cell(row=i,column=1).value,None,None])
        if sheet.cell(row=i,column=1).value==None:
            stop=True
    return numbers[:-1]

# a=get_number(FILE_NAME)